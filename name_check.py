import openpyxl


class NameValidations:

    def __init__(self, file_path, last_row_file, append_file):
        self.file_path = file_path
        self.append_file = append_file
        self.last_row_file = last_row_file
        self.last_row_number = self.load_last_row_number()
        self.r_num = 0  # Initialize r_num as a class attribute

    def __checkValid__(self, name):
        name = str(name)
        valid_name = False
        valid_chars = []
        with open("companies.txt", "r") as f:
            for line in f:
                valid_chars.append(line.strip())
        for index in valid_chars:
            if index in name.lower().split(" "):
                valid_name = True
                return valid_name
        return False

    def load_last_row_number(self):
        try:
            with open(self.last_row_file, "r") as file:
                last_row_number = int(file.read())
            return last_row_number
        except FileNotFoundError:
            return 0

    def get_next_row_data(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        user_name = sheet.cell(row=self.last_row_number + 2, column=1).value
        formal_address = sheet.cell(row=self.last_row_number + 2, column=5).value
        city = sheet.cell(row=self.last_row_number + 2, column=6).value
        state = sheet.cell(row=self.last_row_number + 2, column=7).value

        return user_name, formal_address, city, state

    def save_last_row_number(self, last_row_number):
        with open(self.last_row_file, "w") as file:
            file.write(str(last_row_number))

    def save_value(self, sheet, value):
        sheet.cell(row=self.last_row_number + 2, column=8).value = value
        sheet.cell(row=self.last_row_number + 2, column=12).value = (
            self.last_row_number + 2
        )

    def appending_sheet(self, name, address, city, state):
        wb = openpyxl.load_workbook(self.append_file)
        sheet = wb.active
        sheet.cell(row=self.r_num + 2, column=1).value = name
        sheet.cell(row=self.r_num + 2, column=2).value = address
        sheet.cell(row=self.r_num + 2, column=3).value = city
        sheet.cell(row=self.r_num + 2, column=4).value = state
        sheet.cell(row=self.r_num + 2, column=12).value = self.last_row_number + 2
        wb.save(self.append_file)
        self.r_num += 1  # Increment r_num after appending

    def company_checker(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        while True:
            user_name, formal_address, city, state = self.get_next_row_data()
            if not user_name:
                print("No More Rows To Process")
                break

            if self.__checkValid__(user_name):
                print(f"Company Found: {user_name} is valid.")
                self.save_value(sheet, "It's a Company")
                self.appending_sheet(user_name, formal_address, city, state)

            self.last_row_number += 1
            self.save_last_row_number(self.last_row_number)

        wb.save(self.file_path)


file_path = "companies.xlsx"
last_row_file = "last_row_number.txt"
append_file = "New.xlsx"

if __name__ == "__main__":
    validation_obj = NameValidations(file_path, last_row_file, append_file)
    validation_obj.company_checker()
